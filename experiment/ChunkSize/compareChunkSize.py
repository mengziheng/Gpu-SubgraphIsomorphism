import os
import subprocess
import re
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# 定义文件名列表和文件前缀
file_list_path = "/data/zh_dataset/processed_graph_challenge_dataset/snap"
timeForChunkSize = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/CompareChunkSize/time.xlsx"
speedupForChunkSize = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/CompareChunkSize/speedup.xlsx"
AllChunkSizespeedupForChunkSize = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/CompareChunkSize/all_chunk_size_speedup.xlsx"

N = 50

# 定义参数范围
param_range = range(1,N)
pattern_list = ["Q0","Q1","Q2","Q3","Q6","Q7","Q8","Q11","Q12"]

workbook = openpyxl.Workbook()
sheet = workbook.active

# 标题行列
sheet.cell(row=1, column=1, value="Pattern/ChunkSize")  # 左上角单元格
for i, ChunkSize in enumerate(range(1,N)):
    sheet.cell(row=1, column=i + 2, value=ChunkSize)

for i, Pattern in enumerate(pattern_list):
    sheet.cell(row=i + 2, column=1, value=Pattern)


# 存储测试结果
results = []
i = 0
file_names = os.listdir(file_list_path)

for i,pattern in enumerate(pattern_list):
    ave_results = []
    command = f"./run.sh {pattern}"
    os.system(command)
    print(i,pattern)
    # if(i == 2):
    #     break
    # 遍历每个数据集
    num = 0
    for j,file_name in enumerate(file_names):
        if(j % 2 == 1):
            num = num + 1
            continue
        # 构建文件路径
        file_path = os.path.join(file_list_path, file_name)

        # 存储当前数据集的测试结果
        dataset_results = []
        # 遍历每个参数
        for param in param_range:
            if(j == 0):
                ave_results.append(0)
            # 执行测试并获取结果时间
            command = f"mpirun -n 1 ./subgraphmatch.bin {file_path} {pattern} 1 0.1 4 216 1024 {param}"
            regex_pattern = r"graph : ([\w\-\/]+) time is : (\d+\.\d+) ms,count is : (\d+)"
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
            # 逐行读取输出并提取信息
            for line in iter(process.stdout.readline, b''):
                line = line.decode().strip()
                match = re.search(regex_pattern, line)
                if match:
                    result_time = match.group(2)
                    dataset_results.append(result_time)
        dataset_results = [float(dataset_results[0]) / float(t) for t in dataset_results[0:]]
        print(dataset_results)
        for t in range(len(dataset_results)):
            ave_results[t] += dataset_results[t]
    for t in range(len(dataset_results)):
        ave_results[t] = ave_results[t] / num
        # ave_results[t] = ave_results[t] / 2
        sheet.cell(row=i + 2, column=t + 2, value=ave_results[t])

workbook.save(AllChunkSizespeedupForChunkSize)

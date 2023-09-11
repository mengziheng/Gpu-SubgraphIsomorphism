import os
import subprocess
import re
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

file_list_path = "graph_dataset.txt"
output_path = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/SMOG_scalability/xlsx"
pattern_list = ["Q0", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6"]
# param_range = [1, 2, 4, 8, 16, 32, 64, 128, 256, 512, 1024]
param_range = [1, 2, 4, 8]


results = []
with open(file_list_path, "r") as file:
    file_names = file.read().splitlines()

for k, file_name in enumerate(file_names):
    output_file = os.path.join(output_path, file_name[file_name.rfind("/") + 1 :] + ".xlsx")
    command = f"touch {output_file}"
    os.system(command)
    print(output_file)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, pattern in enumerate(pattern_list):
        print(i, pattern)
        sheet.cell(row=1, column=i + 2, value=pattern)
        dataset_results = []
        for j, param in enumerate(param_range):
            sheet.cell(row=j + 2, column=1, value=param)
            command = f"cd ../../final_version/ && python script.py --input_graph_folder {file_name} --input_pattern {pattern} --N {param}"
            print(command)
            regex_pattern = r"graph : ([\w\-\/]+) time is : (\d+\.\d+) ms,count is : (\d+)"
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
            for line in iter(process.stdout.readline, b""):
                line = line.decode().strip()
                match = re.search(regex_pattern, line)
                if match:
                    result_time = match.group(2)
                    count = match.group(3)
                    results.append(result_time)
                    print(pattern, param, [file_name, result_time, count])
                    sheet.cell(row=j + 2, column=i + 2, value=result_time)
    workbook.save(output_file)

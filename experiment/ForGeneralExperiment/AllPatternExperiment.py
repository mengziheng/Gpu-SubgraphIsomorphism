import os
import subprocess
import re
import openpyxl

pattern_list = ["Q1","Q2","Q3","Q6","Q7","Q8","Q11","Q12"]
workbook = openpyxl.Workbook()
sheet = workbook.active

folder_path = "/data/zh_dataset/processed_graph_challenge_dataset/snap"
# folder_path = sys.argv[1]

output_path = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/AllPattern"
output_file = os.path.join(output_path, folder_path.split('/')[-1] + ".xlsx")

file_names = os.listdir(folder_path)
results = []

parameters = []

# 标题行列
sheet.cell(row=1, column=1, value="Graph/Pattern")  # 左上角单元格
for i, pattern in enumerate(pattern_list):
    sheet.cell(row=1, column=2*i + 2, value=pattern+" time")
    sheet.cell(row=1, column=2*i + 3, value=pattern + " count")

for i, filename in enumerate(file_names):
    sheet.cell(row=i + 2, column=1, value=filename)

for i,pattern in enumerate(pattern_list):
    command = f"./run.sh {pattern}"
    os.system(command)

    # if(i == 2):
    #     break

    for j,file_name in enumerate(file_names):
        if(not (file_name == "friendster" or file_name == "flickrEdges")):
            continue
        print(file_name + " order : " + str(i))
        dir_name = os.path.join(folder_path, file_name)
        command = f"mpirun -n 1 ./subgraphmatch.bin {dir_name} {pattern} 1 0.25 4 216 1024 10"
        regex_pattern = r"graph : ([\w\-\/\.]+) time is : (\d+\.\d+) ms,count is : (\d+)"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)

        for line in iter(process.stdout.readline, b''):
            line = line.decode().strip()
            match = re.search(regex_pattern, line)
            if match:
                graph_name = match.group(1)
                graph_name = graph_name.replace(folder_path,"")
                graph_name = graph_name.replace("/","")
                time = match.group(2)
                count = match.group(3)
                results.append([graph_name, time, count])
                print([graph_name, time, count])
                sheet.cell(row=j+2, column=2*i+2, value=time)
                sheet.cell(row=j+2, column=2*i+3, value=count)

workbook.save(output_file)

import os
import subprocess
import re
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

# 定义文件名列表和文件前缀
file_list_path = "graph_dataset.txt"
output_path = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/Scalability"


pattern_list = ["Q0", "Q1", "Q2", "Q3", "Q6", "Q7", "Q11", "Q12"]
# pattern_list = ["Q2", "Q6", "Q12"]
# pattern_list = ["Q0"]
# 定义参数范围
# param_range = [1,2,4,8]
param_range = [1, 2, 4, 8, 16, 32, 64, 128, 256, 512, 1024]
# param_range = [1, 2, 4]
# 存储测试结果
results = []

# 从文件中逐行读取数据集文件名
with open(file_list_path, "r") as file:
    file_names = file.read().splitlines()

for k, file_name in enumerate(file_names):
    print(output_path)
    print(file_name[file_name.rfind("/") :] + ".xlsx")
    output_file = os.path.join(output_path, file_name[file_name.rfind("/") + 1 :] + ".xlsx")
    command = f"touch {output_file}"
    os.system(command)
    print(output_file)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, pattern in enumerate(pattern_list):
        print(i, pattern)
        command = f"./run.sh {pattern}"
        os.system(command)
        # 遍历每个数据集
        sheet.cell(row=1, column=i + 2, value=pattern)
        # 构建文件路径
        dataset_results = []
        # 遍历每个参数
        for j, param in enumerate(param_range):
            sheet.cell(row=j + 2, column=1, value=param)
            # 执行测试并获取结果时间
            command = f"mpirun -n 2 ./subgraphmatch.bin {file_name} {pattern} {param} 0.25 4 216 1024 1"
            print(command)
            regex_pattern = r"graph : ([\w\-\/]+) time is : (\d+\.\d+) ms,count is : (\d+)"
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
            # 逐行读取输出并提取信息
            for line in iter(process.stdout.readline, b""):
                line = line.decode().strip()
                match = re.search(regex_pattern, line)
                if match:
                    result_time = match.group(2)
                    count = match.group(3)
                    results.append(result_time)
                    print(pattern, param, [file_name, result_time, count])
                    sheet.cell(row=j + 2, column=i + 2, value=result_time)
    workbook.save(output_file)

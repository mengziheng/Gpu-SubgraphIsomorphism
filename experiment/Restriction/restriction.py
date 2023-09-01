import os
import subprocess
import re
import openpyxl

pattern_list = ["Q6"]
# pattern_list = ["Q0","Q1","Q2","Q3","Q6","Q7","Q11","Q12"]
workbook = openpyxl.Workbook()
sheet = workbook.active

folder_path = "/data/zh_dataset/processed_graph_challenge_dataset/Synthetic"
# folder_path = sys.argv[1]

output_file = "/home/zhmeng/GPU/Gpu-SubgraphIsomorphism/result/Restriction/withoutRestriction.xlsx"

file_names = os.listdir(folder_path)
results = []

parameters = []

# 标题行列
sheet.cell(row=1, column=1, value="Graph/Pattern")  # 左上角单元格
for i, pattern in enumerate(pattern_list):
    sheet.cell(row=1, column= i + 2, value=pattern+" time")

    

for i,pattern in enumerate(pattern_list):
    command = f"./run.sh {pattern}"
    os.system(command)
    j = -1
    for file_name in file_names:
        if(not file_name == "P1a"):
            continue
        j = j + 1
        sheet.cell(row=j + 2, column=1, value=file_name)
        print(file_name + " order : " + str(i))
        dir_name = os.path.join(folder_path, file_name)
        command = f"mpirun -n 1 ./subgraphmatch.bin {dir_name} {pattern} 1 0.1 8 216 1024 10"
        regex_pattern = r"graph : ([\w\-\/\.]+) time is : (\d+\.\d+) ms,count is : (\d+)"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)

        for line in iter(process.stdout.readline, b''):
            line = line.decode().strip()
            match = re.search(regex_pattern, line)
            if match:
                graph_name = match.group(1)
                graph_name = graph_name.replace(folder_path,"")
                graph_name = graph_name.replace("/","")
                time = match.group(2)
                count = match.group(3)
                results.append([graph_name, time, count])
                print([graph_name, time, count])
                sheet.cell(row=j+2, column=i+2, value=time)

workbook.save(output_file)

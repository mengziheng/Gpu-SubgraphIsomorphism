# experiment on different pattern

import os
import subprocess
import re
import openpyxl
import pandas as pd

# folder_path = sys.argv[1]

pattern_list = ["Q0","Q1","Q2","Q3","Q4","Q5","Q6"]
folder_path = "../data/processed_graph"
output_path = "../data/result"
output_file = os.path.join(output_path, folder_path.split('/')[-1] + ".xlsx")

file_names = os.listdir(folder_path)
results = []
parameters = []
workbook = openpyxl.Workbook()
sheet = workbook.active


# 标题行列
sheet.cell(row=1, column=1, value="Graph/Pattern")  # 左上角单元格
for i, pattern in enumerate(pattern_list):
    sheet.cell(row=1, column=2*i + 2, value=pattern+" time")
    sheet.cell(row=1, column=2*i + 3, value=pattern + " count")

for i, filename in enumerate(file_names):
    sheet.cell(row=i + 2, column=1, value=filename)

for i,pattern in enumerate(pattern_list):
    for j,file_name in enumerate(file_names):
        if(file_name != "amazon0302"):
            continue
        print(file_name + " order : " + str(i))
        dir_name = os.path.join(folder_path, file_name)
        command = f"cd ../SMOG/ && python script.py --input_graph_folder {dir_name} --input_pattern {pattern}"
        regex_pattern = r"graph : ([\w\-\/\.]+) time is : (\d+\.\d+) ms,count is : (\d+)"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)

        for line in iter(process.stdout.readline, b''):
            line = line.decode().strip()
            match = re.search(regex_pattern, line)
            if match:
                graph_name = match.group(1)
                graph_name = graph_name.replace(folder_path,"")
                graph_name = graph_name.replace("/","")
                time = match.group(2)
                count = match.group(3)
                results.append([graph_name, time, count])
                print([graph_name, time, count])
                sheet.cell(row=j+2, column=2*i+2, value=time)
                sheet.cell(row=j+2, column=2*i+3, value=count)

workbook.save(output_file)
# 读取上一步保存的Excel文件
df = pd.read_excel(output_file, sheet_name="Sheet")
df_value = df.sort_values(by=["Graph/Pattern"], ascending=True)
# 保存文件
writer = pd.ExcelWriter(output_file)
df_value.to_excel(writer, sheet_name='Sheet', index=False)
writer.close()
